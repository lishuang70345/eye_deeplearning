from openpyxl import load_workbook, Workbook
import json
# 表格转换为json
wb = load_workbook("眼科-必有症状--或有症状 - 202105.xlsx")
ws = wb["Sheet1"]

rows = ws.max_row
cols = ws.max_column

print(rows, cols)

disease_info = {}

for row in range(2, rows + 1):

    disease_name = ws.cell(row=row, column=1).value     # 病名
    disease_must = ws.cell(row=row, column=2).value     # 必有症状
    disease_or = ws.cell(row=row, column=3).value       # 或有症状
    disease_history = ws.cell(row=row, column=4).value  # 病史
    disease_age = ws.cell(row=row, column=5).value      # 年龄优先条件
    disease_not = ws.cell(row=row, column=6).value      # 否定
    disease_acute = ws.cell(row=row, column=7).value    # 急慢性

    # print(disease_name, disease_must, disease_or)

    disease_info[disease_name] = {}

    # 获取必有症状
    disease_info[disease_name]['must'] = list(filter(None, disease_must.split("；")))

    # 获取或有症状
    if disease_or:
        disease_info[disease_name]['or'] = list(filter(None, disease_or.split("；")))
    else:
        disease_info[disease_name]['or'] = []

    # 获取病史
    if disease_history:
        disease_info[disease_name]['history'] = list(filter(None, disease_history.split("；")))
    else:
        disease_info[disease_name]['history'] = []

    # 获取年龄优先条件
    if disease_age:
        disease_info[disease_name]['age'] = list(filter(None, disease_age.split("；")))
    else:
        disease_info[disease_name]['age'] = []

    # 获取否定症状
    if disease_not:
        disease_info[disease_name]['not'] = list(filter(None, disease_not.split("；")))
    else:
        disease_info[disease_name]['not'] = []

    # 获取急慢性
    if disease_acute:
        disease_info[disease_name]['acute'] = True
    else:
        disease_info[disease_name]['acute'] = False

for key in disease_info.keys():
    print(key, disease_info[key])

with open("eye_disease_new.json", "w") as fp:
    json.dump(disease_info, fp)
